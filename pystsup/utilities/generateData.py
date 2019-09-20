"""
    An AI Tool for Student-Supervisor Allocation.
    
    Package: pystsup
    Module: utilities
    File: generateData.py
    
    Purpose:  Contains Required functions to read input data from excel file, save results into an excel file
              and create input data excel file.
             
    Author : Rithin Chalumuri
    Version: 1.0 
    Date   : 21/7/17
    
"""


import openpyxl as xl
from openpyxl import *
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.chart import *

from .acmParser import getPath, parseFile
from pystsup.data import Student
from pystsup.data import Supervisor


def getData(stuFile,supFile, keywordsFile="pystsup/test/acm.txt"):
    """
    Function to read input data from excel files and convert it into the required format by GA.

    Parameters:
    
        stuFile (string) - the filepath of the student input data.
        supFile (string) - the filepath of the supervisor input data.

    Returns:

        students (dictionary) - dictionary of all students with their details (id, preferences ,name, realID)
        supervisors (dictionary) - dictionary of all supervisors with their details (id,preferences,quota, name, realID)
    """

    topicNames,topicPaths,topicIDs,levels = parseFile(keywordsFile)

    stuSheet = (xl.load_workbook(stuFile)).active
    supSheet = (xl.load_workbook(supFile)).active

    m = len(list(supSheet.rows))-1
    n = len(list(stuSheet.rows))-1

    numberOfStudents = n
    numberOfSupervisors = m

    stuRows = list(stuSheet.rows)
    del stuRows[0]
    supRows = list(supSheet.rows)
    del supRows[0]

    stuID = "student"
    supID = "supervisor"

    
    students = {}
    supervisors = {}

    #Getting Students
    
    for row in stuRows:
        
        n-=1
        rank = 0
        stu_List = {}
        
        stu = stuID + str(numberOfStudents - n)
        
        realID = row[0].value
        name = row[1].value
        keywords = []
        
        for i in range(2,len(row)):
            val = row[i].value
            val = val.lower().strip()
            if  val != None:
                keywords.append(val)
            else:
                keywords.append(topicIDs[1])
                

        for kw in keywords:
            rank+=1
            stu_List[rank] = [kw,getPath(kw,topicNames, topicPaths, topicIDs)]
        

        studentObject = Student(stu,stu_List,realID,name)
        students[stu] = studentObject


    #Getting Supervisors
    
    for row in supRows:

        m-=1
        rank = 0
        sup_List = {}

        sup = supID + str(numberOfSupervisors - m)

        realID = row[0].value
        name = row[1].value
        quota = int(row[2].value)
        keywords = []
        
        for i in range(3,len(row)):
            val = row[i].value
            val = val.lower().strip()
            if  val != None:
                keywords.append(val)
            else:
                keywords.append(topicIDs[1])
                

        for kw in keywords:
            rank+=1
            sup_List[rank] = [kw,getPath(kw,topicNames, topicPaths, topicIDs)]

        supervisorObject = Supervisor(sup,sup_List,quota,realID,name)
        supervisors[sup] = supervisorObject
        

    return students,supervisors

def scanInputData(stuFile, supFile, keywordsFile):

    topicNames,topicPaths,topicIDs,levels = parseFile(keywordsFile)

    stuWB = xl.load_workbook(stuFile)
    stuSheet = (stuWB).active
    
    supWB = xl.load_workbook(supFile)
    supSheet = (supWB).active

    stuRows = list(stuSheet.rows)
    del stuRows[0]
    supRows = list(supSheet.rows)
    del supRows[0]

    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

    nofill = PatternFill(fill_type='none')

    errorStu = 0
    errorSup = 0

    print("Scanning Student File...")
    for i in range(len(stuRows)):

        row = stuRows[i]

        for j in range(2,len(row)):
            val = row[j].value
            if (val == None):
                continue
            val = val.lower().strip()
            if val not in topicNames:

                #print(f"Found Error in Row {i+1} Keyword {j-2} = {val}")
                errorStu+=1
                stuSheet[f"{chr(65+j)}{i+2}"].fill = redFill
            else:
                stuSheet[f"{chr(65+j)}{i+2}"].fill = nofill
    


    print("\n\nScanning Supervisor Excel File...")

    for i in range(len(supRows)):

        row = supRows[i]

        for j in range(3,len(row)):
            val = row[j].value
            if (val == None):
                continue
            val = val.lower().strip()

            if val not in topicNames:
                #print(f"Found Error in Row {i+1} Keyword {j-2} = {val}")
                errorSup+=1
                supSheet[f"{chr(65+j)}{i+2}"].fill = redFill
            else:
                supSheet[f"{chr(65+j)}{i+2}"].fill = nofill

    
    print("\nScan Complete..\n==========================")
    print(f"Total Errors Found in Student Input File = {errorStu}")
    print(f"Total Errors Found in Supervisor Input File = {errorSup}")


    stuWB.save(stuFile)
    supWB.save(supFile)

    if (errorStu > 0 or errorSup >0):
        print("\nErrors marked and saved in provided excel files\n")
    return errorStu, errorSup


def createExcelFile(filename,data,student):
    """
    Function to create an excel input file for student/supervisor.

    Parameters:
        filename (string) - the name of the file.
        data (list) - list of tuples. Each tuple represents a row, each value of column is spereated by comma in the tuple.
        student (Bool) - True if it is a student input file. Otherwise, False.
    """

    wb = Workbook()
    ws = wb.active
    if student:
        heading = ("Student ID","Student Name","Keyword 1","Keyword 2","Keyword 3","Keyword 4","Keyword 5")
        
    else:
        heading = ("Supervisor ID","Supervisor Name","Quota","Keyword 1","Keyword 2","Keyword 3","Keyword 4","Keyword 5")

    ws.append(heading)

    for i in data:
        ws.append(i)

    wb.save(filename)

    

def writeFrontier(filename, front,metricData, supervisors, students):
    """
    Function to save the final results in an excel file after the completion of the GA.

    Parameters:
    
        filename (string) - the name of the file we want save results in.
        front (list) - list of solutions in the first front.
        metricData (dictionary) - metric data associated with the GA (time taken, max Fitness, diversity etc.)
        supervisors (dictionary) - dictionary of all supervisors with their details (id,preferences,quota, name, realID)
        students (dictionary) - dictionary of all students with their details (id, preferences ,name, realID)
    """

    filename += '.xlsx'
    
    wb = Workbook()

    #Creating the Statistics Sheet
    
    ws = wb.active
    ws.title = "GA Statistics"
    ws.append(("Total Number of Generations",metricData['numberOfGenerations']))
    ws.append(("Total Time Taken (s)",metricData['total_time_generations']))
    ws.append(("Time Per Generation (s)",metricData['avg_time_generation']))
    ws.append((None,None))
    ws.append(("Initial Maximum Student Fitness (Fst)",metricData['initial_maxFst']))
    ws.append(("Final Maximum Student Fitness (Fst)",metricData['final_maxFst']))
    ws.append((None,None))
    ws.append(("Initial Minimum Student Fitness (Fst)",metricData['initial_minFst']))
    ws.append(("Final Minimum Student Fitness (Fst)",metricData['final_minFst']))
    ws.append((None,None))
    ws.append(("Initial Maximum Supervisor Fitness (Fsup)",metricData['initial_maxFsup']))
    ws.append(("Final Maximum Supervisor Fitness (Fsup)",metricData['final_maxFsup']))
    ws.append((None,None))
    ws.append(("Initial Minimum Supervisor Fitness (Fsup)",metricData['initial_minFsup']))
    ws.append(("Final Minimum Supervisor Fitness (Fsup)",metricData['final_minFsup']))
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 17

    
    #Creating the Overview Sheet
    wb.create_sheet("Overview")
    ws = wb["Overview"]
     
    heading = ("Sheet Number","Student Fitness","Supervisor Fitness")
    ws.append(heading)
    
    
    maxFst = max(front, key=lambda x: x.getFst()).getFst()
    maxFsup = max(front, key=lambda x: x.getFsup()).getFsup()
    minFst = min(front,key=lambda x: x.getFst()).getFst()
    minFsup = min(front,key=lambda x: x.getFsup()).getFsup()


    #Filter the frontier with only unique values

    front = set(front)
    n = len(front)
    
    if len(front) > 1:
        front = sorted(front,key=lambda x: (((x.getFsup()-minFsup)/(maxFsup-minFsup))*((x.getFst()-minFst)/(maxFst-minFst))),reverse=True)
        
        
    
    count = 1

    sheetnames = []
    widths = [25,25,15,15]
    letters = ['A','B','C','D']
    
    for sol in front:
        
        sheetName = "Solution " + str(count)
        wb.create_sheet(sheetName)
        sheetnames.append((sheetName,sol.getFst(),sol.getFsup()))
        ws = wb[sheetName]
        

        supEdges = sol.getGraph().getEdges()

        heading = ("Supervisor Name","Student Name","Supervisor ID","Student ID")
        
        ws.append(heading)
        
        for sup in supEdges:
            
            supName = supervisors[sup].getSupervisorName()
            supID = supervisors[sup].getRealID()
            
            for stu in supEdges[sup]:

                stuName = students[stu].getStudentName()
                stuID = students[stu].getRealID()

                tempRow = (supName,stuName,supID,stuID)
                
                ws.append(tempRow)


        for i,width in enumerate(widths):
            ws.column_dimensions[letters[i]].width = width
                

        count+=1


    #Change current worksheet to 'Overview'
        
    ws = wb['Overview']

    #Add the Data into cells
    
    for i in sheetnames:
        ws.append(i)

    ws["A"+str(n+4)]="Diversity Metric"
    ws["B"+str(n+4)] = metricData['diversity']

    #Create the chart and add it to the worksheet
    
    chart = ScatterChart(scatterStyle='marker')
    chart.title = "Pareto Optimal Frontier"
    chart.x_axis.title = 'Student Fitness (Fst)'
    chart.y_axis.title = 'Supervisor Fitness (Fsup)'
    chart.style = 3
    
    xvalues = Reference(ws,min_col=2,min_row=2,max_row=n+1)

    yvalues = Reference(ws,min_col=3,min_row=2,max_row=n+1)

    series = Series(yvalues,xvalues)

    series.marker = xl.chart.marker.Marker('circle')
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    chart.legend = None
    
    ws.add_chart(chart,"A"+str(n+7))

    #Set the column Widths
    
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17

    #Save the File
    
    wb.save(filename)

    
                   
        
